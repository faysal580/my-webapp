"""
Product Image Scraper (web-app version of scrape_images.py)
Reads an uploaded .xlsx catalog, visits each unique product URL, extracts up
to MAX_IMAGES full-size image links, and writes an updated .xlsx.
"""
import re
import json
import time
from pathlib import Path

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

DELAY = 1.5
MAX_RETRIES = 3
TIMEOUT = 15
MAX_IMAGES = 5

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

BROWSER_REQUIRED_DOMAINS = ("othoba.com", "pickaboo.com")
SIZE_RE = re.compile(r"-\d+x\d+(?=\.[a-zA-Z]{2,5}$)")


def strip_size(url):
    return SIZE_RE.sub("", url)


def is_product_image(url):
    low = url.lower()
    bad = ("placeholder", "logo", "icon", "favicon", "banner",
           "pixel", "1x1", "spacer", "woocommerce-placeholder")
    return (
        any(low.endswith(ext) for ext in (".jpg", ".jpeg", ".png", ".webp"))
        and not any(b in low for b in bad)
    )


def find_url_column(df, log):
    for name in ("Product URL", "product_url", "URL", "url", "Link", "link"):
        if name in df.columns:
            return name
    for col in df.columns:
        sample = df[col].dropna().astype(str)
        if sample.str.startswith("http").mean() > 0.5:
            log(f"  Auto-detected URL column: '{col}'")
            return col
    raise ValueError(
        "Could not find a URL column. Make sure one column contains product "
        "page links starting with https://"
    )


# ── Site-specific extractors ────────────────────────────────────────────
def extract_images_inest(soup):
    found, seen = [], set()

    def add(url):
        if not url:
            return
        url = url.strip().replace("//posadmin", "/posadmin")
        if url not in seen and is_product_image(url):
            seen.add(url)
            found.append(url)

    for img in soup.find_all("img"):
        for attr in ("src", "data-src", "data-lazy-src"):
            src = img.get(attr, "") or ""
            if "/posadmin/images/product/large/" in src:
                if src.startswith("//"):
                    src = "https:" + src
                elif src.startswith("/"):
                    src = "https://inest.com.bd" + src
                add(src)
        if len(found) >= MAX_IMAGES:
            break

    if len(found) < MAX_IMAGES:
        og = soup.find("meta", property="og:image")
        if og and og.get("content"):
            add(og["content"])
    return found[:MAX_IMAGES]


def extract_images_rokomari(soup):
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        url = og["content"].strip()
        if is_product_image(url):
            return [url]
    return []


def extract_images_othoba(soup):
    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        url = og["content"].strip()
        if is_product_image(url):
            return [url]
    return []


def extract_images_pickaboo(soup):
    found, seen = [], set()

    def add(url):
        if not url:
            return
        url = url.strip()
        if url.startswith("//"):
            url = "https:" + url
        if url not in seen and is_product_image(url):
            seen.add(url)
            found.append(url)

    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        add(og["content"])

    for img in soup.find_all("img"):
        for attr in ("src", "data-src", "data-lazy-src"):
            src = img.get(attr, "") or ""
            if "pickaboo.com" in src and "/uploads/" in src:
                add(src)
        if len(found) >= MAX_IMAGES:
            break
    return found[:MAX_IMAGES]


def fetch_with_browser(url, log):
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.common.by import By
    except ImportError:
        log("  [!] Selenium not installed. Run:  pip install selenium")
        log("      Also ensure ChromeDriver is in your PATH.")
        return None

    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )

    driver = None
    try:
        driver = webdriver.Chrome(options=options)
        driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        driver.get(url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)
        return driver.page_source
    except Exception as e:
        log(f"  [!] Browser fetch failed: {e}")
        return None
    finally:
        if driver:
            driver.quit()


def extract_images(html, page_url):
    soup = BeautifulSoup(html, "lxml")

    if "inest.com.bd" in page_url:
        return extract_images_inest(soup)
    if "rokomari.com" in page_url:
        return extract_images_rokomari(soup)
    if "othoba.com" in page_url:
        return extract_images_othoba(soup)
    if "pickaboo.com" in page_url:
        return extract_images_pickaboo(soup)

    found, seen = [], set()

    def add(url):
        if not url:
            return
        clean = strip_size(url.strip())
        if clean and clean not in seen and is_product_image(clean):
            seen.add(clean)
            found.append(clean)

    for script in soup.find_all("script"):
        text = script.string or ""
        if not text:
            continue
        for m in re.finditer(
            r'"(?:full_src|full|src|url)"\s*:\s*"(https?://[^"]+\.[a-zA-Z]{2,5})"', text
        ):
            add(m.group(1))
        if len(found) >= MAX_IMAGES:
            break

    for tag in soup.find_all(attrs={"data-product_variations": True}):
        try:
            for v in json.loads(tag["data-product_variations"]):
                img = v.get("image", {})
                for k in ("full_src", "src", "url"):
                    add(img.get(k, ""))
        except (json.JSONDecodeError, TypeError):
            pass

    if len(found) < MAX_IMAGES:
        domain = re.sub(r"https?://([^/]+).*", r"\1", page_url)
        for img in soup.find_all("img"):
            src = img.get("src") or img.get("data-src") or img.get("data-lazy-src") or ""
            if domain in src or "/wp-content/uploads/" in src:
                add(src)
            if len(found) >= MAX_IMAGES:
                break

    if len(found) < MAX_IMAGES:
        for img in soup.select(".woocommerce-product-gallery__image img"):
            add(img.get("src") or img.get("data-src") or "")

    return found[:MAX_IMAGES]


def needs_browser(url):
    return any(domain in url for domain in BROWSER_REQUIRED_DOMAINS)


def fetch_with_retry(session, url, log):
    if needs_browser(url):
        log("  (using browser fetch)")
        return fetch_with_browser(url, log)

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = session.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            return r.text
        except requests.RequestException as e:
            log(f"    attempt {attempt}/{MAX_RETRIES} failed: {e}")
            if attempt < MAX_RETRIES:
                time.sleep(DELAY * 2)
    return None


def parse_pasted_urls(urls_text):
    """Splits a textarea blob into a clean list of unique-order URLs."""
    if not urls_text:
        return []
    lines = re.split(r"[\r\n]+", urls_text)
    urls = []
    for line in lines:
        # allow comma/space separated URLs on the same line too
        for piece in re.split(r"[,\s]+", line.strip()):
            piece = piece.strip()
            if piece:
                urls.append(piece)
    return urls


def run(output_dir: Path, log, xlsx_file: Path = None, urls_text: str = None):
    """Main entry point called by the web app. Returns the output file path.

    Accepts an uploaded .xlsx catalog (xlsx_file), pasted URLs (urls_text),
    or both — pasted URLs are appended as extra rows when a file is also
    supplied.
    """
    pasted_urls = parse_pasted_urls(urls_text)

    if xlsx_file:
        input_path = Path(xlsx_file)
        log(f"Input  : {input_path.name}")

        df = pd.read_excel(input_path)
        url_col = find_url_column(df, log)
        wb = load_workbook(input_path)
        ws = wb.active
        url_col_idx = df.columns.get_loc(url_col) + 1
        output_path = Path(output_dir) / (input_path.stem + "_updated.xlsx")

        if pasted_urls:
            log(f"Appending {len(pasted_urls)} pasted URL(s) to the catalog…")
            next_row = ws.max_row + 1
            for u in pasted_urls:
                ws.cell(row=next_row, column=url_col_idx, value=u)
                next_row += 1
            extra_df = pd.DataFrame({url_col: pasted_urls})
            df = pd.concat([df, extra_df], ignore_index=True)
    else:
        if not pasted_urls:
            raise ValueError(
                "Please upload an .xlsx file or paste at least one product URL."
            )
        log(f"Input  : {len(pasted_urls)} pasted URL(s) (no file uploaded)")

        url_col = "Product URL"
        df = pd.DataFrame({url_col: pasted_urls})
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value=url_col)
        for i, u in enumerate(pasted_urls, start=2):
            ws.cell(row=i, column=1, value=u)
        url_col_idx = 1
        output_path = Path(output_dir) / "pasted_urls_updated.xlsx"

    log(f"Output : {output_path.name}")

    unique_urls = df[url_col].dropna().unique().tolist()
    cache = {}

    session = requests.Session()
    session.headers.update(HEADERS)

    log(f"Scraping {len(unique_urls)} unique product URLs …")

    for i, url in enumerate(unique_urls, 1):
        log(f"[{i}/{len(unique_urls)}] {url}")
        html = fetch_with_retry(session, url, log)
        if html:
            imgs = extract_images(html, url)
            cache[url] = imgs
            log(f"  -> {len(imgs)} image(s) found")
        else:
            cache[url] = []
            log("  -> failed; leaving blank")
        if i < len(unique_urls):
            time.sleep(DELAY)

    log("Writing output …")
    img_start_col = ws.max_column + 1

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="left", vertical="center")

    for j in range(1, MAX_IMAGES + 1):
        col = img_start_col + j - 1
        cell = ws.cell(row=1, column=col, value=f"Image Link {j}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = 60

    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=url_col_idx).value
        imgs = cache.get(str(url).strip(), [])
        for j, img_url in enumerate(imgs):
            cell = ws.cell(row=row_idx, column=img_start_col + j, value=img_url)
            cell.alignment = cell_align

    wb.save(output_path)

    total_rows = len(df)
    total_unique = len(unique_urls)
    total_found = sum(1 for u in unique_urls if cache.get(u))
    total_missing = total_unique - total_found

    log("")
    log(f"Done! Saved -> {output_path.name}")
    log(f"  Rows total   : {total_rows}")
    log(f"  URLs scraped : {total_unique}")
    log(f"  Found        : {total_found}")
    log(f"  Not found    : {total_missing}")

    return [output_path]
